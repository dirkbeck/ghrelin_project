import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import binom

plt.rcParams['pdf.fonttype'] = 42
plt.rcParams['axes.spines.top'] = False
plt.rcParams['axes.spines.right'] = False

def p_success(x):
    return 1 / (1 + np.exp(x))

# params
max_dim = 10
x0 = np.linspace(-2, 2, 400)
p = p_success(x0)

# compute pmfs
probs = [binom.pmf(k, max_dim, p) for k in range(max_dim+1)]
probs = [np.maximum(arr, 0) for arr in probs]
total = np.sum(probs, axis=0)
probs = [arr/total for arr in probs]

# colors
cmap = plt.get_cmap('viridis')
colors = cmap(np.linspace(0, 1, max_dim+1))

# plot
fig, (ax0, ax1) = plt.subplots(1, 2, figsize=(10, 4))

# left: DMS
x_plot = x0 + 2
ax0.stackplot(x_plot, *probs, colors=colors, alpha=0.85)
ax0.set_xlim(0, 4)
ax0.set_ylim(0, 1)
ax0.set_xlabel('DMS striosome activity')
ax0.set_ylabel('Probability')

#DLS (mirrored)
x_plot_m = -x0 + 2
ax1.stackplot(x_plot_m, *probs, colors=colors, alpha=0.85)
ax1.set_xlim(0, 4)
ax1.set_ylim(0, 1)
ax1.set_xlabel('DLS striosome activity')

# colorbar
sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(0, max_dim))
cb = plt.colorbar(sm, ax=[ax0, ax1])
cb.set_label('decision-space dimensionality')

plt.tight_layout()
plt.savefig('decision_space.pdf')
plt.show()


# ── SOURCE DATA EXPORT ──────────────────────────────────────────────────────

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

TITLE = "decision_space"
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{TITLE} Source Data.xlsx")

# Downsample to ~100 pts
idx = np.round(np.linspace(0, len(x0) - 1, 100)).astype(int)
x_dms_ds = x_plot[idx]
x_dls_ds = x_plot_m[idx]
probs_ds  = [arr[idx] for arr in probs]

# ── helpers ──────────────────────────────────────────────────────────────────

DATA_FILL  = PatternFill("solid", start_color="D6E4F0")
PARAM_FILL = PatternFill("solid", start_color="D9EAD3")

def write_header(ws, row, cols, fill):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def write_row(ws, row, values):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(name="Arial")

def write_params(ws, start_row, params):
    write_header(ws, start_row, ["Parameter", "Value", "Description"], PARAM_FILL)
    for i, (name, val, desc) in enumerate(params, 1):
        write_row(ws, start_row + i, [name, val, desc])

def autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 60)

# ── shared parameters ────────────────────────────────────────────────────────

SHARED_PARAMS = [
    ("max_dim",       max_dim,    "Maximum decision-space dimensionality (n in Binomial)"),
    ("x0_min",        -2,         "Min of latent sigmoid input x0"),
    ("x0_max",         2,         "Max of latent sigmoid input x0"),
    ("n_x0_pts",       400,       "Number of x0 points before downsampling"),
    ("n_ds_pts",       100,       "Number of points after downsampling for export"),
    ("sigmoid",       "1/(1+exp(x))", "p_success formula (maps x0 → Binomial p)"),
    ("k_values",      f"0 to {max_dim}", "Outcome values k for Binom(n=max_dim, p)"),
    ("normalization", "row-wise sum-to-1", "PMFs renormalized so sum over k = 1 at each x0"),
    ("alpha",          0.85,      "Stackplot fill alpha"),
    ("cmap",          "viridis",  "Matplotlib colormap; k=0→purple, k=max_dim→yellow"),
    ("cmap_norm",     f"Normalize(vmin=0, vmax={max_dim})", "Colormap normalization range"),
    ("x_shift_DMS",   "+2",       "DMS x-axis = x0 + 2 (shifts range to [0,4])"),
    ("x_shift_DLS",   "negate then +2", "DLS x-axis = -x0 + 2 (mirrors DMS)"),
    ("xlim",          "0–4",      "x-axis display range for both panels"),
    ("ylim",          "0–1",      "y-axis display range (cumulative probability)"),
]

data_header = ["x (striosome activity)"] + [f"P(k={k})" for k in range(max_dim + 1)]

# ── build workbook ───────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─── Fig5D ───────────────────────────────────────────────────────────────────

ws5d = wb.create_sheet("Fig5D")
write_header(ws5d, 1, data_header, DATA_FILL)
for i, xi in enumerate(x_dms_ds):
    write_row(ws5d, i + 2, [round(float(xi), 6)] + [round(float(probs_ds[k][i]), 8) for k in range(max_dim + 1)])

param_row_5d = len(x_dms_ds) + 3
write_params(ws5d, param_row_5d, SHARED_PARAMS)
autosize(ws5d)

# ─── Fig S14I ────────────────────────────────────────────────────────────────

ws14i = wb.create_sheet("Fig S14I")

# DMS block
ws14i.cell(row=1, column=1, value="Panel: DMS (left)").font = Font(name="Arial", bold=True, italic=True)
write_header(ws14i, 2, data_header, DATA_FILL)
for i, xi in enumerate(x_dms_ds):
    write_row(ws14i, i + 3, [round(float(xi), 6)] + [round(float(probs_ds[k][i]), 8) for k in range(max_dim + 1)])

# DLS block
dls_start = len(x_dms_ds) + 5
ws14i.cell(row=dls_start - 1, column=1, value="Panel: DLS (right, mirrored)").font = Font(name="Arial", bold=True, italic=True)
write_header(ws14i, dls_start, data_header, DATA_FILL)
for i, xi in enumerate(x_dls_ds):
    write_row(ws14i, dls_start + i + 1, [round(float(xi), 6)] + [round(float(probs_ds[k][i]), 8) for k in range(max_dim + 1)])

# shared params block
param_row_14i = dls_start + len(x_dls_ds) + 3
write_params(ws14i, param_row_14i, SHARED_PARAMS)
autosize(ws14i)

# ── save ─────────────────────────────────────────────────────────────────────

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")