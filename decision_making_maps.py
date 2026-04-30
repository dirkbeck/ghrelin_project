import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import TwoSlopeNorm, LinearSegmentedColormap

plt.rcParams['pdf.fonttype'] = 42
np.random.seed(0)

gamma, w_r, w_p, alpha, beta, k_sig = 2.0, 1.0, 1.0, 5.0, 5.0, 10.0
thr_low, thr_high, D_mid, k_dim = 1.0, 10.0, 5.0, 1.0

I_r = np.linspace(0, 1, 200)
I_p = np.linspace(0, 1, 200)
Ir, Ip = np.meshgrid(I_r, I_p)


def sigmoid(x, k=1.0):
    return 1.0 / (1.0 + np.exp(-k * x))


def w_dims(D):
    return 1.0 - sigmoid(D - D_mid, k=k_dim)


def compute_P_avoid(D, use_dims=True):
    Q_avoid = w_p * Ip
    Q_approach = w_r * Ir
    if use_dims:
        g3, g4 = sigmoid(D - thr_low, k=k_sig), sigmoid(D - thr_high, k=k_sig)
        fade = w_dims(D)
        Q_avoid += fade * alpha * g3 * (Ir * Ip)
        Q_approach += fade * beta * g4 * (Ir * (1 - Ip))
    num = np.exp(gamma * Q_avoid)
    return num / (num + np.exp(gamma * Q_approach))


P_base = compute_P_avoid(0.0, True)
P_low_dims, P_low_nodims = compute_P_avoid(1.0, True), compute_P_avoid(1.0, False)
P_high_dims, P_high_nodims = compute_P_avoid(10.0, True), compute_P_avoid(10.0, False)

dP_low, dP_low_nd = P_low_dims - P_base, P_low_nodims - P_base
dP_high, dP_high_nd = P_high_dims - P_base, P_high_nodims - P_base

cmap_rw = LinearSegmentedColormap.from_list("approach_avoid",
                                            [(0.0, "#1B7837"),
                                             (0.4, "white"),
                                             (0.6, "white"),
                                             (1.0, "#762A83")])

norm_P  = TwoSlopeNorm(vmin=0.0, vcenter=0.5, vmax=1.0)
norm_dP = TwoSlopeNorm(vmin=-0.5, vcenter=0.0, vmax=0.5)


def plot_panels(axs, P_base, P_D, dP, label, use_dims, row):
    mats = [(P_base, f'P(avoid) base\n(dims={"on" if use_dims else "off"})', norm_P),
            (P_D,   f'P(avoid) {label}\n(dims={"on" if use_dims else "off"})', norm_P),
            (dP,    f'ΔP(avoid) {label}\n(dims={"on" if use_dims else "off"})', norm_dP)]

    for col, (mat, title, norm) in enumerate(mats):
        axs[row, col].imshow(mat, origin='lower', extent=(0, 1, 0, 1), cmap=cmap_rw, norm=norm)
        axs[row, col].set_title(title)
        axs[row, col].set_xlabel('Reward (I_r)')
        axs[row, col].set_ylabel('Cost (I_p)')
        axs[row, col].grid(alpha=0.3)


# low ghrelin
fig1, axs1 = plt.subplots(2, 3, figsize=(9, 6), constrained_layout=True)
plot_panels(axs1, P_base, P_low_dims,  dP_low,    'low-ghrelin', True,  0)
plot_panels(axs1, P_base, P_low_nodims, dP_low_nd, 'low-ghrelin', False, 1)
fig1.colorbar(plt.cm.ScalarMappable(norm=norm_P,  cmap=cmap_rw),
              ax=axs1[:, 0:2], label='P(avoid)  [green=approach, purple=avoid]')
fig1.colorbar(plt.cm.ScalarMappable(norm=norm_dP, cmap=cmap_rw),
              ax=axs1[:, 2],   label='ΔP(avoid)  [green=↑approach, purple=↑avoid]')
plt.savefig('low_ghrelin_maps.pdf', dpi=300)
plt.show()

# high ghrelin
fig2, axs2 = plt.subplots(2, 3, figsize=(9, 6), constrained_layout=True)
plot_panels(axs2, P_base, P_high_dims,  dP_high,    'high-ghrelin', True,  0)
plot_panels(axs2, P_base, P_high_nodims, dP_high_nd, 'high-ghrelin', False, 1)
fig2.colorbar(plt.cm.ScalarMappable(norm=norm_P,  cmap=cmap_rw),
              ax=axs2[:, 0:2], label='P(avoid)  [green=approach, purple=avoid]')
fig2.colorbar(plt.cm.ScalarMappable(norm=norm_dP, cmap=cmap_rw),
              ax=axs2[:, 2],   label='ΔP(avoid)  [green=↑approach, purple=↑avoid]')
plt.suptitle('High‐Ghrelin Decision Maps\n(green=higher approach, purple=higher avoidance)')
plt.savefig('high_ghrelin_maps.pdf', dpi=300)
plt.show()

# ─────────────────────────────────────────────
# SOURCE DATA EXPORT
# ─────────────────────────────────────────────
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TITLE = "Ghrelin Decision Maps"
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{TITLE} Source Data.xlsx")

N_GRID = 50
COL_GAP = 1
PANEL_COLS = 3
PANEL_STRIDE = PANEL_COLS + COL_GAP

FONT_NAME  = "Arial"
BLUE_FILL  = PatternFill("solid", start_color="BDD7EE")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")


def downsample_2d(mat, n=N_GRID):
    idx = np.round(np.linspace(0, mat.shape[0] - 1, n)).astype(int)
    return mat[np.ix_(idx, idx)], np.linspace(0, 1, n)


def style_data_hdr(cell):
    cell.font = Font(name=FONT_NAME, bold=True)
    cell.fill = BLUE_FILL
    cell.alignment = Alignment(horizontal="center")


def style_param_hdr(cell):
    cell.font = Font(name=FONT_NAME, bold=True)
    cell.fill = GREEN_FILL
    cell.alignment = Alignment(horizontal="left")


def write_panel(ws, mat, ax_vals, col_start, row_label_row, label, data_start_row):
    title_cell = ws.cell(row=row_label_row, column=col_start, value=label)
    title_cell.font = Font(name=FONT_NAME, bold=True, italic=True)

    for offset, h in enumerate(["Cost (I_p)", "Reward (I_r)", "Value"]):
        c = ws.cell(row=data_start_row, column=col_start + offset, value=h)
        style_data_hdr(c)

    r = data_start_row + 1
    n = len(ax_vals)
    for i in range(n):
        for j in range(n):
            ws.cell(row=r, column=col_start,     value=round(float(ax_vals[i]), 4)).font = Font(name=FONT_NAME)
            ws.cell(row=r, column=col_start + 1, value=round(float(ax_vals[j]), 4)).font = Font(name=FONT_NAME)
            ws.cell(row=r, column=col_start + 2, value=round(float(mat[i, j]), 6)).font  = Font(name=FONT_NAME)
            r += 1
    return r


def write_params_block(ws, params, row_start, col_start=1):
    hdr_row = row_start + 1
    for offset, h in enumerate(["Parameter", "Value", "Description"]):
        c = ws.cell(row=hdr_row, column=col_start + offset, value=h)
        style_param_hdr(c)
    for i, (name, val, desc) in enumerate(params):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=col_start,     value=name).font = Font(name=FONT_NAME)
        ws.cell(row=r, column=col_start + 1, value=val).font  = Font(name=FONT_NAME)
        ws.cell(row=r, column=col_start + 2, value=desc).font = Font(name=FONT_NAME)


def autosize(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 30)


SHARED_PARAMS = [
    ("gamma",    gamma,    "Softmax inverse temperature for action selection"),
    ("w_r",      w_r,      "Weight on reward Q-value (Q_approach)"),
    ("w_p",      w_p,      "Weight on cost Q-value (Q_avoid)"),
    ("alpha",    alpha,    "Interaction gain for Q_avoid (dims-on only)"),
    ("beta",     beta,     "Interaction gain for Q_approach (dims-on only)"),
    ("k_sig",    k_sig,    "Sigmoid steepness for gating functions g3, g4"),
    ("thr_low",  thr_low,  "D threshold: g3 = sigmoid(D - thr_low, k_sig)"),
    ("thr_high", thr_high, "D threshold: g4 = sigmoid(D - thr_high, k_sig)"),
    ("D_mid",    D_mid,    "D midpoint for dimensional-weighting sigmoid w_dims"),
    ("k_dim",    k_dim,    "Sigmoid steepness for w_dims"),
    ("grid_pts", 200,      "Full simulation grid points per axis"),
    ("export_grid_pts", N_GRID, "Downsampled grid points per axis in this export"),
    ("I_r range", "0–1",  "Reward information axis range"),
    ("I_p range", "0–1",  "Cost/punishment information axis range"),
    ("Colormap",  "approach_avoid: #1B7837 (green) → white (0.4–0.6) → #762A83 (purple)", "LinearSegmentedColormap"),
    ("norm_P",   "TwoSlopeNorm vmin=0, vcenter=0.5, vmax=1",   "Normalization for P(avoid) panels"),
    ("norm_dP",  "TwoSlopeNorm vmin=-0.5, vcenter=0, vmax=0.5","Normalization for ΔP(avoid) panels"),
    ("P(avoid)", "exp(γ·Q_avoid) / (exp(γ·Q_avoid) + exp(γ·Q_approach))", "Softmax avoidance probability"),
    ("Q_avoid (dims-on)",    "w_p·I_p + w_dims(D)·α·g3·(I_r·I_p)",      "g3=sigmoid(D-thr_low,k_sig)"),
    ("Q_approach (dims-on)", "w_r·I_r + w_dims(D)·β·g4·(I_r·(1-I_p))", "g4=sigmoid(D-thr_high,k_sig)"),
    ("Q_avoid (dims-off)",   "w_p·I_p",   "No interaction term"),
    ("Q_approach (dims-off)","w_r·I_r",   "No interaction term"),
    ("w_dims(D)", "1 - sigmoid(D - D_mid, k_dim)", "Interaction fade as D increases"),
    ("Color: green (#1B7837)", "P(avoid)=0", "High approach probability"),
    ("Color: purple (#762A83)", "P(avoid)=1", "High avoidance probability"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)


def populate_sheet(ws, panels, extra_params):
    LABEL_ROW = 1
    DATA_HDR_ROW = 2
    DATA_START = DATA_HDR_ROW

    max_end_row = DATA_START
    for p_idx, (label, mat, _norm, _desc) in enumerate(panels):
        col_start = 1 + p_idx * PANEL_STRIDE
        ds_mat, ax_vals = downsample_2d(mat)
        end_row = write_panel(ws, ds_mat, ax_vals,
                              col_start=col_start,
                              row_label_row=LABEL_ROW,
                              label=label,
                              data_start_row=DATA_START)
        max_end_row = max(max_end_row, end_row)

    params_all = SHARED_PARAMS + extra_params
    write_params_block(ws, params_all, row_start=max_end_row, col_start=1)
    autosize(ws)


ws5E = wb.create_sheet("Fig5E")
panels_5E = [
    ("P(avoid) base | D=0, dims-on",         P_base,     norm_P,  ""),
    ("P(avoid) low-ghrelin | D=1, dims-on",  P_low_dims, norm_P,  ""),
    ("ΔP(avoid) low-ghrelin | D=1, dims-on", dP_low,     norm_dP, ""),
]
extra_5E = [
    ("D_base", 0.0, "Ghrelin drive for baseline surface"),
    ("D_low",  1.0, "Ghrelin drive for low-ghrelin condition"),
    ("dims",  "on", "Dimensional weighting active for this figure row"),
]
populate_sheet(ws5E, panels_5E, extra_5E)

wsS10B = wb.create_sheet("SFig. S10B")
panels_S10B = [
    ("P(avoid) base | D=0, dims-on",          P_base,      norm_P,  ""),
    ("P(avoid) high-ghrelin | D=10, dims-on", P_high_dims, norm_P,  ""),
    ("ΔP(avoid) high-ghrelin | D=10, dims-on", dP_high,    norm_dP, ""),
]
extra_S10B = [
    ("D_base", 0.0,  "Ghrelin drive for baseline surface"),
    ("D_high", 10.0, "Ghrelin drive for high-ghrelin condition"),
    ("dims",  "on",  "Dimensional weighting active for this figure row"),
]
populate_sheet(wsS10B, panels_S10B, extra_S10B)

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")