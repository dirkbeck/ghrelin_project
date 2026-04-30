import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
plt.rcParams['pdf.fonttype'] = 42

inverse_temp = 1.5

def sigmoid(x, k=1.0):
    return 1.0 / (1.0 + np.exp(-k * x))

def striosome_activity(dose, peak_dose=3.0, width=1.2):
    return np.exp(-((dose - peak_dose) ** 2) / (2 * width ** 2))

n = 200
contexts = np.linspace(0, 1, n)
doses    = np.linspace(0, 10, n)

ctx_col  = contexts.reshape(n, 1)
dose_row = doses.reshape(1, n)

Q_base_col = 3.0 * (ctx_col - 0.5)

strio_row = striosome_activity(dose_row)

conflict_sigma = 0.40
conflict_weight_col = np.exp(
    -((ctx_col - 0.5) ** 2) / (2 * conflict_sigma ** 2)
)

ghrelin_strength = 2.0
ghrelin_effect = -ghrelin_strength * conflict_weight_col * strio_row

Q_total    = Q_base_col + ghrelin_effect
P_approach = sigmoid(Q_total, k=inverse_temp)

cmap_rw = LinearSegmentedColormap.from_list("approach_avoid",
                                            [(0.0, "#1B7837"),
                                             (0.4, "white"),
                                             (0.6, "white"),
                                             (1.0, "#762A83")])

fig1, ax1 = plt.subplots(figsize=(8, 5))
im = ax1.imshow(P_approach, origin='lower', aspect='auto',
                extent=[doses.min(), doses.max(),
                        contexts.min(), contexts.max()],
                cmap=cmap_rw, vmin=0, vmax=1)
ax1.contour(doses, contexts, P_approach,
            levels=[0.5], colors='black', linewidths=2)
ax1.set_yticks([0.0, 0.5, 1.0])
ax1.set_yticklabels(['Cost Only', 'Conflict', 'Reward Only'])
ax1.set_ylabel('Task Context')
ax1.set_xlabel('Ghrelin Dose (arb. u.)')
ax1.set_title('Phase Diagram: IBU modifies behavior based on context and dose')
plt.colorbar(im, ax=ax1, label='P(Approach)')
plt.tight_layout()
plt.savefig('phase_plot.pdf')
plt.show()

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TITLE = "SFig. 10C"
PANEL = "SFig. 10C"
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        f"{TITLE} Source Data.xlsx")

DS = 50
ds_idx      = np.linspace(0, n - 1, DS, dtype=int)
ds_doses    = doses[ds_idx]
ds_contexts = contexts[ds_idx]
ds_P        = P_approach[np.ix_(ds_idx, ds_idx)]

HDR_DATA   = {"font": Font(name="Arial", bold=True, color="FFFFFF"),
               "fill": PatternFill("solid", start_color="2E75B6"),
               "align": Alignment(horizontal="center")}
HDR_PARAMS = {"font": Font(name="Arial", bold=True, color="FFFFFF"),
               "fill": PatternFill("solid", start_color="375623"),
               "align": Alignment(horizontal="center")}
CELL_FONT  = Font(name="Arial")

def style_row(ws, row_idx, style):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.font      = style["font"]
            cell.fill      = style["fill"]
            cell.alignment = style["align"]

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = PANEL

ws.append(["Context \\ Dose (arb. u.)"] + [f"{d:.4f}" for d in ds_doses])
style_row(ws, 1, HDR_DATA)

for i, ctx_val in enumerate(ds_contexts):
    row = [f"{ctx_val:.4f}"] + [float(f"{v:.6f}") for v in ds_P[i, :]]
    ws.append(row)

ws.append([])
ws.append(["Note: P(Approach) = 0.5 contour marks the decision boundary (plotted in black)."])
ws.append([])
ws.append([])

param_row = ws.max_row + 1
ws.append(["Parameter", "Value", "Description"])
style_row(ws, param_row, HDR_PARAMS)

params = [
    ("inverse_temp",            inverse_temp,      "Softmax inverse temperature (k) for sigmoid"),
    ("sigmoid_formula",         "1/(1+exp(-k·Q))", "Choice probability formula"),
    ("peak_dose",               3.0,               "Striosome tuning curve peak (arb. u.)"),
    ("width",                   1.2,               "Striosome tuning curve SD"),
    ("striosome_formula",       "exp(-((dose-peak)^2)/(2*width^2))", "Striosome activity g(dose)"),
    ("Q_base_scale",            3.0,               "Slope of context-dependent baseline Q"),
    ("Q_base_formula",          "3*(ctx - 0.5)",   "Context baseline Q value"),
    ("conflict_sigma",          conflict_sigma,    "Conflict zone width (Gaussian SD in context)"),
    ("conflict_weight_formula", "exp(-((ctx-0.5)^2)/(2*conflict_sigma^2))", "Conflict weight h(ctx)"),
    ("ghrelin_strength",        ghrelin_strength,  "Ghrelin modulation magnitude"),
    ("ghrelin_effect_formula",  "-ghrelin_strength * h(ctx) * g(dose)", "Ghrelin suppression term"),
    ("Q_total_formula",         "Q_base + ghrelin_effect", "Net decision variable"),
    ("n_ctx_points",            n,                 "Grid resolution (context axis)"),
    ("n_dose_points",           n,                 "Grid resolution (dose axis)"),
    ("ctx_range",               "0 to 1",          "Context axis range"),
    ("dose_range",              "0 to 10",         "Dose axis range (arb. u.)"),
    ("colormap",                "approach_avoid",  "LinearSegmentedColormap: #1B7837→white→white→#762A83"),
    ("cmap_stops",              "0.0, 0.4, 0.6, 1.0", "Colormap breakpoint positions"),
    ("cmap_colors",             "#1B7837, white, white, #762A83", "Green→White→Purple (colorblind-safe)"),
    ("vmin",                    0.0,               "Colormap normalization minimum"),
    ("vmax",                    1.0,               "Colormap normalization maximum"),
    ("contour_level",           0.5,               "Contour line drawn at P(Approach)=0.5"),
    ("export_grid_size",        f"{DS}×{DS}",      "Downsampled grid dimensions for this export"),
]

for p in params:
    ws.append(list(p))

for row in ws.iter_rows():
    for cell in row:
        if not cell.font.bold:
            cell.font = CELL_FONT

auto_width(ws)
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")